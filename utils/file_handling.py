from openpyxl import load_workbook


def is_valid_excel_file(path, required_sheets, min_threshold=0.7):
    try:
        wb = load_workbook(path, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
        found = sum(1 for name in required_sheets if name in sheetnames)
        return (found / len(required_sheets)) >= min_threshold
    except Exception:
        return False
