from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import filedialog, messagebox
import tkinter as tk
from PIL import Image, ImageTk
import os
from openpyxl import load_workbook

REQUIRED_SHEETS = [
    "APPENDIX C-C1 FAS VER template",
    "ULC Coverpage",
    "32.5 Response Times",
    "32.6 Large Scale Network System",
    "32.11",
    "32.12",
    "32.13",
    "ULC Cover Page",
    "Deficiency Summary",
    "EXT only",
    "ELU only",
    "HOSES only",
    "20.1 | Report",
    "20.2 | Deficiencies",
    "20.3 | Recommendations",
    "21 | Documentation",
    "29",
    "30",
    "31 Documentation (2)",
    "22.1 | CU or Transp Insp",
    "32 ControlUnit|Transponder (2)",
    "22.2 | CU or Transp Test",
    "22.3 + 22.4 | Voice & PS",
    "32.7",
    "32.8 Power Supply (2)",
    "22.5 | Power Supply(s)",
    "22.6 | Annunciator(s)",
    "22.7 | Annun & Seq Disp",
    "22.9 + 22.10 | Printer",
    "23.1 Field Device Legend",
    "23.2 Device Record",
    "23.3 CircuitFaultTolerance",
]

ICON_PATH = "./assets/excel_icon.png"
TILE_WIDTH = 110
TILE_HEIGHT = 100
TILES_PER_ROW = 4


def is_valid_excel_file(path, required_sheets, threshold=0.7):
    try:
        wb = load_workbook(path, read_only=True)
        present_sheets = set(wb.sheetnames)
        required = set(required_sheets)
        found_count = len(required & present_sheets)
        return (found_count / len(required)) >= threshold
    except Exception:
        return False


class V8MergerApp(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("V8 Merger")
        self.geometry("600x600")
        self.configure(bg="white")
        self.resizable(False, False)

        self.selected_files = []
        self.file_tiles = []
        self.icon_image = None
        self.create_widgets()

    def create_widgets(self):
        self.instruction_label = tk.Label(
            self,
            text="Drag and drop V8 Excel inspection files below or click Browse",
            bg="white",
            font=("Segoe UI", 11, "bold"),
            fg="#333",
        )
        self.instruction_label.pack(pady=(20, 2))

        self.sub_instruction_label = tk.Label(
            self,
            text="Only V8-format fire inspection files are supported.",
            bg="white",
            font=("Segoe UI", 9, "italic"),
            fg="#6c757d",
        )
        self.sub_instruction_label.pack(pady=(0, 10))

        self.drop_frame = tk.Frame(self, bg="#f8f9fa", bd=2, relief="ridge", height=120)
        self.drop_frame.pack(padx=20, fill=tk.X)
        self.drop_frame.pack_propagate(False)

        self.drop_label = tk.Label(
            self.drop_frame,
            text="Drop Excel files here",
            bg="#f8f9fa",
            font=("Segoe UI", 10, "italic"),
            fg="#495057",
        )
        self.drop_label.pack(expand=True)

        self.drop_frame.drop_target_register(DND_FILES)
        self.drop_frame.dnd_bind("<<Drop>>", self.handle_drop)

        self.browse_button = tk.Button(
            self,
            text="Browse",
            command=self.browse_files,
            font=("Segoe UI", 10),
            width=25,
            bg="#2c7be5",
            fg="white",
            relief="raised",
            bd=2,
        )
        self.browse_button.pack(pady=(10, 10))

        self.grid_frame = tk.Frame(self, bg="white")
        self.grid_frame.pack(padx=20, pady=(10, 5))

        self.status_label = tk.Label(
            self, text="", bg="white", font=("Segoe UI", 9), fg="gray"
        )
        self.status_label.pack(pady=5)

        self.clear_button = tk.Button(
            self,
            text="Clear All",
            command=self.clear_all_files,
            font=("Segoe UI", 10),
            width=25,
            bg="#6c757d",
            fg="white",
            relief="raised",
            bd=2,
        )
        self.clear_button.pack(pady=(5, 15))

        if os.path.exists(ICON_PATH):
            img = Image.open(ICON_PATH).resize((32, 32))
            self.icon_image = ImageTk.PhotoImage(img)

    def browse_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
        self.add_files(files)

    def handle_drop(self, event):
        files = self.tk.splitlist(event.data)
        self.add_files(files)

    def add_files(self, files):
        for f in files:
            cleaned = f.strip('"')
            if cleaned in self.selected_files:
                self.status_label.config(
                    text=f"File Error: Duplicate file '{os.path.basename(cleaned)}'",
                    fg="red",
                )
                return
            elif cleaned.endswith(".xlsx"):
                if is_valid_excel_file(cleaned, REQUIRED_SHEETS):
                    self.selected_files.append(cleaned)
                    self.add_file_tile(cleaned)
                else:
                    self.status_label.config(
                        text=(
                            "File Error: Missing required sheets in "
                            f"'{os.path.basename(cleaned)}'"
                        ),
                        fg="red",
                    )
                    return
            else:
                self.status_label.config(
                    text=(
                        "File Error: Not a valid Excel file → "
                        f"{os.path.basename(cleaned)}"
                    ),
                    fg="red",
                )
                return

        self.status_label.config(
            text=f"{len(self.selected_files)} files selected", fg="gray"
        )

    def add_file_tile(self, filepath):
        idx = len(self.file_tiles)
        row = idx // TILES_PER_ROW
        col = idx % TILES_PER_ROW

        tile = tk.Frame(
            self.grid_frame,
            width=TILE_WIDTH,
            height=TILE_HEIGHT,
            bg="#f1f3f5",
            bd=1,
            relief="solid",
        )
        tile.grid(row=row, column=col, padx=10, pady=10)
        tile.pack_propagate(False)

        if self.icon_image:
            icon_label = tk.Label(tile, image=self.icon_image, bg="#f1f3f5")
            icon_label.pack(pady=(5, 0))

        label = tk.Label(
            tile,
            text=os.path.basename(filepath),
            bg="#f1f3f5",
            wraplength=90,
            justify="center",
            font=("Segoe UI", 8),
        )
        label.pack(pady=(2, 2))

        remove_btn = tk.Button(
            tile,
            text="✕",
            command=lambda: self.remove_tile(tile, filepath),
            font=("Segoe UI", 8),
            bg="#dc3545",
            fg="white",
            bd=0,
            padx=2,
            pady=0,
        )
        remove_btn.place(relx=1.0, rely=0.0, anchor="ne")

        self.file_tiles.append(tile)

    def remove_tile(self, tile, filepath):
        tile.destroy()
        self.selected_files = [f for f in self.selected_files if f != filepath]
        self.file_tiles.remove(tile)
        self.relayout_tiles()
        self.status_label.config(
            text=f"{len(self.selected_files)} files selected", fg="gray"
        )

    def relayout_tiles(self):
        for idx, tile in enumerate(self.file_tiles):
            row = idx // TILES_PER_ROW
            col = idx % TILES_PER_ROW
            tile.grid(row=row, column=col, padx=10, pady=10)

    def clear_all_files(self):
        if messagebox.askyesno(
            "Clear All Files", "Are you sure you want to remove all selected files?"
        ):
            for tile in self.file_tiles:
                tile.destroy()
            self.selected_files.clear()
            self.file_tiles.clear()
            self.status_label.config(text="All files cleared", fg="gray")


def launch_app():
    app = V8MergerApp()
    app.mainloop()
