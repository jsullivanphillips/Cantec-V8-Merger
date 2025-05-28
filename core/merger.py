from openpyxl import load_workbook, Workbook
from config.sheet_definitions import get_merge_handler


def merge(file_paths):
    input_wbs = [load_workbook(path, data_only=True) for path in file_paths]
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # remove default sheet

    all_sheet_names = set()
    for wb in input_wbs:
        all_sheet_names.update(wb.sheetnames)

    for sheet_name in sorted(all_sheet_names):
        handler = get_merge_handler(sheet_name)
        if handler:
            print(f"Merging sheet: {sheet_name}")
            ws_list = [
                wb[sheet_name] for wb in input_wbs if sheet_name in wb.sheetnames
            ]
            output_ws = output_wb.create_sheet(title=sheet_name)
            handler(ws_list, output_ws)

    return output_wb
